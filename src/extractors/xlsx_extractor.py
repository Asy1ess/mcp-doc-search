"""Excel XLSX text extraction."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from src.extractors._utils import make_document


def extract_xlsx(path: Path) -> ExtractedDocument:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_count = len(workbook.sheetnames)
    sheets: list[str] = []

    try:
        for sheet in workbook.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append("\t".join(cells))
            if rows:
                sheets.append(f"## {sheet.title}\n" + "\n".join(rows))
    finally:
        workbook.close()

    return make_document(
        path,
        "\n\n".join(sheets),
        format="xlsx",
        sheet_count=str(sheet_count),
    )
